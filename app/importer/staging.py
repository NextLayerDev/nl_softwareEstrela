"""Etapa de STAGING: lê a planilha bruta preservando aba, linha e valor literal.

Nenhuma transformação destrutiva aqui — só leitura. O parser consome estas células.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


@dataclass
class CelulaBruta:
    """Uma linha bruta da planilha, indexada por coluna (0-based)."""

    aba: str
    linha: int  # 1-based, como aparece no Excel
    valores: list[object | None] = field(default_factory=list)

    def valor(self, idx: int) -> object | None:
        if idx is None or idx < 0 or idx >= len(self.valores):
            return None
        return self.valores[idx]

    @property
    def vazia(self) -> bool:
        return all(v is None or (isinstance(v, str) and not v.strip()) for v in self.valores)


def ler_staging(caminho: str | Path, abas: list[str]) -> dict[str, list[CelulaBruta]]:
    """Lê as abas pedidas e devolve {aba: [CelulaBruta, ...]} com valores literais.

    Usa `data_only=True` para pegar o valor calculado das células (não a fórmula).
    """
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        resultado: dict[str, list[CelulaBruta]] = {}
        disponiveis = set(wb.sheetnames)
        for aba in abas:
            if aba not in disponiveis:
                # aba esperada ausente: registra vazia para o chamador decidir
                resultado[aba] = []
                continue
            ws = wb[aba]
            linhas: list[CelulaBruta] = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                linhas.append(CelulaBruta(aba=aba, linha=i, valores=list(row)))
            resultado[aba] = linhas
        return resultado
    finally:
        wb.close()
